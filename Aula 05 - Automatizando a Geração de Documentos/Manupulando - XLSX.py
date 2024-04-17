# Instando a biblioteca openpyxl
# pip install openpyxl

'''from openpyxl import load_workbook #from para buscar a bablioteca específica, senão via biscar a biblioteca completa
arquivo = 'planilha.xlsx'
workbook = load_workbook(arquivo)
planilha = workbook.active #ativa para fazer a leitura do openpyxl

for linha in planilha.iter_rows():  #iterar
    for celula in linha:
        print(celula.value, end = '\t')
    print()
'''

#OU]

'''
from openpyxl import load_workbook #from para buscar a bablioteca específica, senão via biscar a biblioteca completa
arquivo = 'planilha.xlsx'
workbook = load_workbook(arquivo)
planilha = workbook.active #ativa para fazer a leitura do openpyxl

for linha in planilha.iter_rows():  #iterar
    for celula in linha:
        print(str(celula.value).ljust(20),end='') #ljust para ajustar as colunas
    print()
'''



'''Exibir uma informação usando referência de celula'''

import openpyxl

planilha = openpyxl.load_workbook('planilha.xlsx')
aba = planilha['PRODUTOS']
print(aba['a2'].value)
print(aba['b2'].value)
